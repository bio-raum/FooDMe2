#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import json


parser = argparse.ArgumentParser(description="Script options")
parser.add_argument("--results")
parser.add_argument("--output")
args = parser.parse_args()


def main(results_in, output):
    with open(results_in) as fi:
        results = json.load(fi)

    # Create an empty workbook and start page
    wb = Workbook()
    ws = wb.active

    ws.title = "FooDMe2 benchmarking"

    ft = Font(name="Sans", bold=True)
    cb_even = PatternFill(fill_type = "solid", fgColor="d9e1f2")
    cb_uneven = PatternFill(fill_type = "solid", fgColor="cdd1d9")
    redfill = PatternFill(fill_type = "solid", fgColor="fc4c4c")
    
    # Track cell positions
    row = 0

    ws.append(["Predicted taxon", "Expected taxon", "Matching rank", "Predicted %", "Expected %", "Classification"])

    for r in ws["A1:F1"]:
        for cell in r:
            cell.font = ft
    
    this_sample = ""
    sample_counter = 0

    # Iterate over each sample and get infos
    for sample in results.keys():

        if sample != this_sample:
            this_sample = sample
            sample_counter += 1

        if sample_counter & 1:
            bgcolor = cb_uneven
        else:
            bgcolor = cb_even

        hits = results[sample]
        for hit in hits:
            row += 1
            ws.append([
                hit["prediction_name"],
                hit["expect_name"],
                hit["match_rank"],
                round(float(hit["pred_freq"]),4)*100,
                round(float(hit["expect_freq"]),4)*100,
                hit["result"].upper()
            ])
            
            ws["A"+str(ws._current_row)].fill = bgcolor
            ws["B"+str(ws._current_row)].fill = bgcolor
            ws["C"+str(ws._current_row)].fill = bgcolor
            ws["D"+str(ws._current_row)].fill = bgcolor
            ws["E"+str(ws._current_row)].fill = bgcolor
            if hit["result"].upper() != "TP":
                ws["F"+str(ws._current_row)].fill = redfill
            else:
                ws["F"+str(ws._current_row)].fill = bgcolor

    # Auto-width for columns
    dim_holder = DimensionHolder(worksheet=ws)

    for column in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(column)] = ColumnDimension(ws, min=column, max=column, width=20)

    ws.column_dimensions = dim_holder

    ws.freeze_panes = ws["A2"]
    wb.save(output)

if __name__ == '__main__':
    main(args.results, args.output)
