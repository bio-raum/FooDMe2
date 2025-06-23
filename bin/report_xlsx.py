#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import glob
import argparse
import json
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


parser = argparse.ArgumentParser(description="Script options")
parser.add_argument("--output")
args = parser.parse_args()


def main(output):
    wb = Workbook()
    ft = Font(name="Sans", bold=True)
    cb_even = PatternFill(fill_type="solid", fgColor="d9e1f2")
    cb_uneven = PatternFill(fill_type="solid", fgColor="cdd1d9")

    # First sheet is the results per sample ================================
    # Get all the TSV files in this directory
    reports = sorted(glob.glob("*.json"))
    bucket_compo = {}  # holds the composition info per sample
    bucket_consensus = {}  # holds the consensus info per sample

    # Read the summary json and get the required data (composition, consensus) per sample
    for report in reports:
        with open(report) as fd:
            jdata = json.load(fd)

        sample = jdata["sample"]
        compo = jdata["composition"]
        consensus = jdata["consensus"]

        bucket_compo[sample] = compo
        bucket_consensus[sample] = consensus

    # start page
    ws = wb.active
    ws.title = "FooDMe2 results"

    # Track cell positions
    row = 0
    ws.append(["Sample", "Taxon", "Percentage", "Reads", "Cluster IDs"])
    for r in ws["A1:E1"]:
        for cell in r:
            cell.font = ft
    this_sample = ""
    sample_counter = 0

    # Iterate over each sample and get taxonomy assignments
    for sample in bucket_compo:
        if sample != this_sample:
            this_sample = sample
            sample_counter += 1
        if sample_counter & 1:
            bgcolor = cb_uneven
        else:
            bgcolor = cb_even

        hits = bucket_compo[sample]
        for hit in hits:
            row += 1
            name = hit["name"]
            perc = round(float(hit["proportion"]), 4) * 100
            reads = hit["reads"]
            cluster_ids = hit["cluster_ids"]

            ws.append([sample, name, perc, reads, cluster_ids])

            for col in ["A", "B", "C", "D", "E"]:
                ws[col + str(ws._current_row)].fill = bgcolor

    # Auto-width for columns
    dim_holder = DimensionHolder(worksheet=ws)
    for column in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(column)] = ColumnDimension(ws, min=column, max=column, width=20)
    ws.column_dimensions = dim_holder
    ws.freeze_panes = ws["A2"]

    # Second sheet is the support for each sequence =========================
    # Start a new sheet
    ws2 = wb.create_sheet("Call Support")

    # Get all the JSON files in this directory
    reports = sorted(glob.glob("*.json"))

    # Track cell positions
    row = 0
    ws2.append(["Sample", "Cluster ID", "Size", "Taxon", "Ranks", "Taxid", "Support [%]", "Consensus"])
    for r in ws2["A1:H1"]:
        for cell in r:
            cell.font = ft
    this_sample = ""
    sample_counter = 0

    # Iterate over each sample
    for sample in bucket_consensus:
        if sample != this_sample:
            this_sample = sample
            sample_counter += 1
        if sample_counter & 1:
            bgcolor = cb_uneven
        else:
            bgcolor = cb_even

        clusters = bucket_consensus[sample]
        # iterate over each cluster and report each one
        # Start with the actual call
        for cluster in clusters:
            id = cluster["query"]
            size = cluster["size"]
            name = cluster["name"]
            rank = cluster["rank"]
            calltaxid = cluster["taxid"]
            support = round(cluster["support"] * 100, 2)
            call = True

            row += 1
            ws2.append([sample, id, size, name, rank, calltaxid, support, call])
            for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                ws2[col + str(ws2._current_row)].fill = bgcolor

            # then go through the tax_list, but skip the taxid that was called
            for hit in cluster["tax_list"]:
                id = cluster["query"]
                size = cluster["size"]
                name = hit["name"]
                rank = "species"  # Always species !
                taxid = hit["taxid"]
                support = round(hit["freq"] * 100, 2)
                call = False  # Always False

                if taxid != calltaxid:
                    row += 1
                    ws2.append([sample, id, size, name, rank, calltaxid, support, call])
                    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                        ws2[col + str(ws2._current_row)].fill = bgcolor

    # Auto-width for columns
    dim_holder = DimensionHolder(worksheet=ws2)
    for column in range(ws2.min_column, ws2.max_column + 1):
        dim_holder[get_column_letter(column)] = ColumnDimension(ws2, min=column, max=column, width=20)
    ws2.column_dimensions = dim_holder
    ws2.freeze_panes = ws["A2"]

    # Write excel file
    wb.save(output)


if __name__ == '__main__':
    main(args.output)
